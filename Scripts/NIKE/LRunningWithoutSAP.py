from openpyxl import Workbook
from openpyxl.styles import Border, Font, Alignment, Side, PatternFill
from datetime import datetime
import win32com.client as win32

olApp = win32.Dispatch('Outlook.Application')
olNS = olApp.GetNameSpace('MAPI')
mail = olApp.CreateItem(0)

f = open('lr.txt', 'r')

cNow = str(datetime.now()).split(' ')[0].split('-')
cDate = cNow[2]
cMonth = cNow[1]
month = datetime.strptime(cMonth, "%m").strftime("%b")
month_names = ('Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec')

wb = Workbook()
ws = wb.active

ws['A1'], ws['B1'] = 'GVID|BPID|SupplierID', 'TimeStamp'
ws['A1'].font = Font(bold=True, size=12)
ws['B1'].font = Font(bold=True, size=12)
ws.column_dimensions['A'].width, ws.column_dimensions['B'].width = 40, 35

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

for line in f:
    if line[:3] in month_names:
        data = line.split('VendorPortal_ProjectVendorPortal_Project:')
        date = data[0].split('\t')[1]
        ids = data[1].split('\t')[1]
        ws.append([ids, date])
else:
    for i in range(1, ws.max_row+1):
        for j in range(1, ws.max_column+1):
            cell_obj = ws.cell(row=i, column=j)
            if i == 1:
                cell_obj.fill = PatternFill(start_color='BEBEBE', end_color='BEBEBE', fill_type='solid')
            cell_obj.alignment = Alignment(horizontal='center', vertical='bottom')
            cell_obj.border = thin_border
    name = f'LongRunning_{month}{cDate}.xlsx'
    path = 'C:\\Users\\aroy11\\Desktop\\Task\\longRunning\\'
    wb.save(path+name)

    mail.Subject = 'Terminate  - Long Running instances in SOA Production'
    mail.Body = """Hi Team,

We have long running instances in SOA1PRD2. Kindly create ITASK  to terminate these instances.
Please find attached the instance details for your reference.

Thanks & Regards,
Arbin Roy
    """
    mail.to = 'Lst-Frontline.FUSION@nike.com'
    mail.Cc = 'Nithya.Kannan@nike.com; Karthick.Kuppan@nike.com; Lst-EOA.Platform.IntegrationOps.Fusion@nike.com'
    mail.Attachments.Add(path+name)
    f.close()
    mail.Display()
