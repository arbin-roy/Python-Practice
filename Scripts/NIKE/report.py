from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, Alignment, Side, PatternFill
from datetime import datetime
f = open('er.txt', encoding='utf8')

cNow = str(datetime.now()).split(' ')[0].split('-')
cDate = cNow[2]
cMonth = cNow[1]
month_name = datetime.strptime(cMonth, "%m").strftime("%b")

wb = Workbook()
ws = wb.active
ws.title = f'{cDate}-{month_name}-{cNow[0]}'

headers = ['Vendor Type', 'Service Name', 'BusinessKey Name', 'BusinessKey Value', 'Time Stamp', 'Reason', 'Solution Scope', 'Error Description']
spliter = 'GVID|SupplierID|BusinessProcessID'
ws.append(['Error Details'])
ws.append(headers)

ws.column_dimensions['A'].width = 45
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 45
ws.column_dimensions['D'].width = 55
ws.column_dimensions['E'].width = 40
ws.column_dimensions['F'].width = 65
ws.column_dimensions['G'].width = 18
ws.column_dimensions['H'].width = 200

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
reasons = {409: '409 error received from Aravo Acknowledgment',
          'Conflicting receive': 'Similar Instance is already running. Need to terminate before reprocessing'}

for line in f:
    if spliter in line and line[0].isalpha():
        mainData = line.split(spliter)
        set1, set2 = mainData[0].split('\t'), mainData[1].split('\t')
        serviceName, timeStamp, ids, error, reason = set1[2], set1[9], set2[1], set2[2], ''
        solutionScope = set1[1] if 'Fusion' == set1[1] else 'Aravo'
        if '409' in error:
            reason = reasons.get(409, 'Null')
        elif 'Conflicting receive' in error:
            reason = reasons.get('Conflicting receive', 'Null')
        else:
            reason = 'Null'
        ws.append(['Indirect', serviceName, spliter, ids, timeStamp, reason, solutionScope, error])
else:
    ws.append(['Terminated Long Running Instances'])
    ws.append([headers[1], headers[2], headers[3]])

    lr, data = load_workbook('longRunning/'+f'LongRunning_{month_name}{cDate}.xlsx'), []
    sheet = lr.active
    for i in range(2, sheet.max_row + 1):
        for j in range(1, sheet.max_column):
            cell = sheet.cell(row=i, column=j)
            data.append(cell.value)
        ws.append([spliter, data[0], data[1]])
        data.clear()
    lr.close()

    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            cell = ws.cell(row=i, column=j)
            if cell.value in headers:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color='BEBEBE', end_color='BEBEBE', fill_type='solid')
            elif i == 1 or cell.value == 'Terminated Long Running Instances':
                cell.font = Font(bold=True, size=12, color='FFFFFF')
                cell.fill = PatternFill(start_color='0D0D0D', end_color='0D0D0D', fill_type='solid')
            elif cell.value:
                cell.border = thin_border
    name = f'VendorPortal_{month_name}{cDate}_Fusion_Error_Report.xlsx'
    wb.save('C:\\Users\\aroy11\\Desktop\\Task\\errorReports\\'+name)
    wb.close()
    f.close()
