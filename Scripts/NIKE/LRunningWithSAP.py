from openpyxl import Workbook
from openpyxl.styles import Border, Font, Alignment, Side, PatternFill
from datetime import datetime
f = open('lr.txt', 'r')

cNow = str(datetime.now()).split(' ')[0].split('-')
cDate = cNow[2]
cMonth = cNow[1]
month = datetime.strptime(cMonth, "%m").strftime("%b")
month_names = ('Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec')

wb = Workbook()
ws = wb.active

ws['A1'], ws['B1'], ws['C1'] = 'GVID|BPID|SupplierID', 'TimeStamp', 'SAP Instance'
ws['A1'].font = Font(bold=True, size=12)
ws['B1'].font = Font(bold=True, size=12)
ws['C1'].font = Font(bold=True, size=12)
ws.column_dimensions['A'].width, ws.column_dimensions['B'].width, ws.column_dimensions['C'].width = 40, 35, 25

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
instance = {1: 'PRD', 2: 'PRA', 3: 'CRP'}

for line in f:
    #fVal = line[0]
    #print(line[:3], month_name)
    #if line[:3] == month_name:
    if line[:3] in month_names:
    #if not fVal.isdigit() and fVal != 'V' and fVal != 'R' and fVal != ' ':
        data = line.split('VendorPortal_ProjectVendorPortal_Project:')
        date = data[0].split('\t')[1]
        ids = data[1].split('\t')[1]
        print(f'Enter SAP Instance for: {ids}')
        ws.append([ids, date, instance[int(input())]])
else:
    for i in range(1, ws.max_row+1):
        for j in range(1, ws.max_column+1):
            cell_obj = ws.cell(row=i, column=j)
            if i == 1:
                cell_obj.fill = PatternFill(start_color='BEBEBE', end_color='BEBEBE', fill_type='solid')
            cell_obj.alignment = Alignment(horizontal='center', vertical='bottom')
            cell_obj.border = thin_border
    name = f'LongRunning_{month}{cDate}.xlsx'
    wb.save('C:\\Users\\aroy11\\Desktop\\Task\\longRunning\\'+name)
    f.close()
