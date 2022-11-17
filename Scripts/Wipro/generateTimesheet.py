import os
from openpyxl import load_workbook
from openpyxl.styles import Border, Font, Alignment, Side, PatternFill
from datetime import datetime
import calendar

datesList, workingHours, totalWorkingHours = [], [], 0
highlighter = ['Sat', 'Sun', 'Leave']
destinationFolder = os.path.dirname(__file__) + '\\doc\\'
fileName = 'Arbin-TS.xlsx'
anyLeaves = [int(i) for i in input('Have you taken any leaves in this month?:  ').split(',')]
dtNow = datetime.now()
monthName = datetime.strptime(str(dtNow.month), "%m").strftime("%b")
wb = load_workbook(f"{destinationFolder}{fileName}")

if 'Sheet1' in wb.sheetnames:
    wb.remove(wb['Sheet1'])

ws = wb.create_sheet(monthName + str(dtNow.year)[-2:])
dayCountOfMonth = calendar.monthrange(dtNow.year, dtNow.month)[1]
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
ws.column_dimensions['A'].width, ws.column_dimensions['B'].width = 14, 15

for i in range(1, dayCountOfMonth + 1):
    datesList.append(f'{i}-{monthName}')
    intWeekDay = calendar.weekday(dtNow.year, dtNow.month, i)
    if intWeekDay < 5 and i not in anyLeaves:
        workingHours.append(8)
        totalWorkingHours += 8
    else:
        workingHours.append(calendar.day_name[calendar.weekday(dtNow.year, dtNow.month, i)][:3] if i not in anyLeaves else 'Leave')
else:
    ws.append(['Employee ID', 'Name', 'Dates'] + datesList + ['Total'])
    ws.append(['40083867', 'Arbin Roy', 'Efforts'] + workingHours + [totalWorkingHours])
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            cell_obj = ws.cell(row=i, column=j)
            if i == 1:
                cell_obj.font = Font(bold=True, size=12)
                cell_obj.fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
            elif cell_obj.value in highlighter:
                cell_obj.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
            cell_obj.alignment = Alignment(horizontal='center', vertical='bottom')
            cell_obj.border = thin_border

wb.save(f"{destinationFolder}{fileName}")
wb.close()
